"""
analise.py
Real Estate Analytics - Versão com Negrito e Padrão Brasileiro
Autora: Vanusagomes-dev
"""
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ARQUIVO = "Analise_Imoveis.xlsx"

def analisar():

    dados = {
        "Valor Venda": [259000.0],
        "Aluguel Bruto": [2200.0],
        "Condomínio": [400.0],
        "IPTU": [80.0],
        "Taxa Administração": [150.0],
        "Provisão": [50.0],
        "Aluguel Líquido": [2050.0],
        "Yield Bruto": [0.1019],
        "Yield Líquido": [0.0949],
        "Rentabilidade": [0.07]
    }

    df = pd.DataFrame(dados)

    df.to_excel(ARQUIVO, index=False)

    formatar_excel()

    print("\n" + "=" * 70)
    print("RELATÓRIO DE ANÁLISE IMOBILIÁRIA")
    print("=" * 70)
    print(df.to_string(index=False))
    print("=" * 70)

    try:
        os.startfile(ARQUIVO)

    except Exception:
        print("Não foi possível abrir a planilha automaticamente.")

def formatar_excel():

    wb = load_workbook(ARQUIVO)
    ws = wb.active

    azul = "1F4E78"

    for celula in ws[1]:
        celula.font = Font(
            color="FFFFFF",
            bold=True
        )

        celula.fill = PatternFill(
            "solid",
            fgColor=azul
        )

    moeda_br = 'R$ #,##0.00_-'
    perc_br = '0.00%'

    # CORREÇÃO DAS COLUNAS
    header_cells = {
        cell.value: cell.column
        for cell in ws[1]
    }
    colunas_moeda = [
        "Valor Venda",
        "Aluguel Bruto",
        "Condomínio",
        "IPTU",
        "Taxa Administração",
        "Provisão",
        "Aluguel Líquido"
    ]

    colunas_percentuais = [
        "Yield Bruto",
        "Yield Líquido",
        "Rentabilidade"
    ]
    for linha in ws.iter_rows(min_row=2):

        for celula in linha:
            celula.font = Font(
                bold=True
            )

        numero_linha = linha[0].row

        for coluna in colunas_moeda:

            if coluna in header_cells:

                ws.cell(
                    row=numero_linha,
                    column=header_cells[coluna]
                ).number_format = moeda_br

        for coluna in colunas_percentuais:

            if coluna in header_cells:

                ws.cell(
                    row=numero_linha,
                    column=header_cells[coluna]
                ).number_format = perc_br
    for coluna in ws.columns:

        letra = get_column_letter(
            coluna[0].column
        )
        largura = max(
            len(str(c.value))
            if c.value else 0
            for c in coluna
        ) + 5

        ws.column_dimensions[letra].width = largura

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(ARQUIVO)

if __name__ == "__main__":
    analisar()

